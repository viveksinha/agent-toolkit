#!/usr/bin/env python3
"""
Extract holdings from a Trendlyne multigroup .xlsx export.
Usage: python3 extract_portfolio.py <xlsx_path>
Outputs: JSON lines (one per holding) + total portfolio value header
"""
import json
import sys
import openpyxl

def extract(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Data Downloader']
    headers = [c.value for c in ws[1]]

    holdings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            d = {h: row[i] for i, h in enumerate(headers) if row[i] is not None}
            holdings.append(d)

    total = sum(h.get('Current Value', 0) or 0 for h in holdings)
    equity = [h for h in holdings if h.get('PE TTM Price to Earnings') is not None]

    print(f"TOTAL_PORTFOLIO: {total:.0f}")
    print(f"EQUITY_COUNT: {len(equity)}")
    print(f"TOTAL_COUNT: {len(holdings)}")
    print("=== ALL HOLDINGS ===")
    for h in holdings:
        print(json.dumps(h, default=str))

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: extract_portfolio.py <path_to_xlsx>", file=sys.stderr)
        sys.exit(1)
    extract(sys.argv[1])
